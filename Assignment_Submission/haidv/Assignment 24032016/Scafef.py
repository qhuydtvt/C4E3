from bs4 import BeautifulSoup

file=open('Scafef.html',"rb")
webpage=file.read()
webpage_string=webpage.decode("utf-8")
soup=BeautifulSoup(webpage_string,"html.parser")
input_tag=soup.find_all('input',value=4)#tim tag

#
# id_tag2=soup.find_all('div',"DuLieu_ChartBCTC")
# print(id_tag2)
# for i in id_tag2:
#     _anh=i.find_all("img")
#     for k in _anh:
#         print(k)
#         print("----------")
# div=soup.find_all("img",{'class':"img_bctc_VNM"})
tds=soup.find("table",id="tableContent")
u1=tds.find_all("tr")
import openpyxl
wb = openpyxl.Workbook()
ws = wb.create_sheet()
ws.title = 'cafef'

i=1
j=1
for k in u1:
    h=k.find_all("td")
    j=1
    for m in h:

        if len(m.text.strip())> 0:
           # print(m.text.strip(), end="")
           # print(" | ", end="")
           ws.cell(row=i,column=j).value = m.text.strip()
           K=ws.cell(row=i,column=j).value
           print(K)
           wb.save(filename=r'D:\book2.xlsx')
           j=j+1
    i=i+1

